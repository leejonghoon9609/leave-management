from http.server import BaseHTTPRequestHandler
import json, io, urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

def thin(color='FFCCCCCC'):
    s = Side(style='thin', color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def med(color='FF888888'):
    s = Side(style='medium', color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def fill(color):
    return PatternFill('solid', fgColor=color)

def font(bold=False, sz=11, color='FF000000'):
    return Font(bold=bold, size=sz, color=color)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

GREEN   = 'FF1D9E75'; GREEN_D  = 'FF0F6E56'; GREEN_L  = 'FFE1F5EE'
BLUE    = 'FF378ADD'; BLUE_D   = 'FF185FA5'; BLUE_L   = 'FFE6F1FB'
RED     = 'FFE24B4A'; RED_D    = 'FFA32D2D'; RED_L    = 'FFFCEBEB'
PURPLE  = 'FF7F77DD'; PURPLE_D = 'FF534AB7'; PURPLE_L = 'FFEEEDFE'
AMBER_D = 'FF854F0B'; AMBER_L  = 'FFFAEEDA'
HEADER  = 'FF2E7D52'; WHITE    = 'FFFFFFFF'
GRAY_L  = 'FFF3F9F6'; NAVY    = 'FF1E3A5F'

def make_excel(payload):
    data      = payload['data']       # [{name, dept, wd, p, l, a, lv, ot, lateRate}, ...]
    daily     = payload['daily']      # [{name, dept, month, rows:[{date,gubun,type,time,status,memo}]}]
    title     = payload['title']      # "2026년 5월 팀 근태 통계(사업관리(현장))"
    month_label = payload['monthLabel']  # "2026년 5월"

    wb = Workbook()
    wb.remove(wb.active)

    # ── 시트1: 월별현황 (막대그래프 + 팀통계 나란히) ─────────────────────
    ws = wb.create_sheet('월별현황')

    BAR_MAX = 20
    max_val = max((max(r['p'], r['l'], r['a'], r['ot']) for r in data), default=1) or 1

    # 열 너비 설정
    ws.column_dimensions['A'].width = 9      # 항목명
    for ci in range(2, BAR_MAX + 2):         # B~U = 막대칸
        ws.column_dimensions[get_column_letter(ci)].width = 1.6
    ws.column_dimensions[get_column_letter(BAR_MAX + 2)].width = 5   # V=합계
    ws.column_dimensions[get_column_letter(BAR_MAX + 3)].width = 2   # W=구분
    # Y~AG = 팀통계
    stat_cols = [8, 14, 6, 8, 6, 6, 6, 6, 7]
    for i, w in enumerate(stat_cols):
        ws.column_dimensions[get_column_letter(BAR_MAX + 4 + i)].width = w

    # 행 높이
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    # 제목 (A1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=BAR_MAX+2)
    c = ws.cell(1, 1, title + ' - 그래프')
    c.font = font(True, 12, HEADER[2:]); c.alignment = align('left')

    # 팀통계 제목 (Y1)
    sc = BAR_MAX + 4
    ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc+8)
    c = ws.cell(1, sc, title)
    c.font = font(True, 12, HEADER[2:]); c.alignment = align('left')

    # 범례 (2행)
    legends = [('■ 정상출근', GREEN), ('■ 지각', BLUE), ('■ 결근', RED), ('■ 특근', PURPLE)]
    for i, (label, color_) in enumerate(legends):
        c = ws.cell(2, 1 + i*5, label)
        c.font = font(True, 10, color_[2:]); c.alignment = align('left')
        c.border = Border(bottom=Side(style='thin', color='FFCCCCCC'))

    # 팀통계 헤더 (Y2)
    stat_hdrs = ['이름', '부서', '평일수', '정상출근', '지각', '결근', '연차', '특근', '지각률']
    for i, h in enumerate(stat_hdrs):
        c = ws.cell(2, sc + i, h)
        c.font = font(True, 10, 'FFFFFFFF')
        c.fill = fill(HEADER[2:])
        c.alignment = align('center')
        c.border = thin('FFAAAAAA')

    bar_defs = [
        ('정상출근', GREEN, GREEN_L),
        ('지각',     BLUE,  BLUE_L),
        ('결근',     RED,   RED_L),
        ('특근',     PURPLE,PURPLE_L),
    ]
    keys = ['p', 'l', 'a', 'ot']

    row = 3
    for ri, emp in enumerate(data):
        # 직원 이름 헤더
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=BAR_MAX+2)
        c = ws.cell(row, 1, emp['name'])
        c.font = font(True, 12, NAVY[2:]); c.alignment = align('left', 'center')
        ws.row_dimensions[row].height = 18
        row += 1

        # 일수 헤더행
        ws.row_dimensions[row].height = 14
        c = ws.cell(row, 1, '일수')
        c.font = font(True, 9, NAVY[2:]); c.alignment = align('center')
        c.border = Border(bottom=Side(style='thin', color='FFCCCCCC'), right=Side(style='thin', color='FFCCCCCC'))
        for d in range(1, BAR_MAX+1):
            c = ws.cell(row, 1+d, d)
            c.font = font(False, 9); c.alignment = align('center')
            c.border = Border(bottom=Side(style='thin', color='FFCCCCCC'))
        c = ws.cell(row, BAR_MAX+2, '합계')
        c.font = font(False, 9); c.alignment = align('center')
        c.border = Border(bottom=Side(style='thin', color='FFCCCCCC'), left=Side(style='thin', color='FFCCCCCC'))
        row += 1

        # 막대 4행
        for ki, (label, col, col_l) in enumerate(bar_defs):
            ws.row_dimensions[row].height = 16
            val = emp[keys[ki]]
            bar_len = round((val / max_val) * BAR_MAX) if max_val > 0 else 0

            c = ws.cell(row, 1, label)
            c.font = font(False, 10, col[2:]); c.alignment = align('right', 'center')
            c.border = Border(bottom=Side(style='thin',color='FFCCCCCC'), right=Side(style='thin',color='FFCCCCCC'))

            for bi in range(1, BAR_MAX+1):
                c = ws.cell(row, 1+bi, '')
                c.fill = fill(col[2:] if bi <= bar_len else 'FFFAFAFA')
                c.border = Border(
                    bottom=Side(style='thin', color='FFCCCCCC'),
                    right=Side(style='thin', color='FFCCCCCC') if bi == BAR_MAX else None
                )

            c = ws.cell(row, BAR_MAX+2, val)
            c.font = font(True, 11, col[2:]); c.alignment = align('center')
            c.border = Border(bottom=Side(style='thin',color='FFCCCCCC'), left=Side(style='thin',color='FFCCCCCC'))
            row += 1

        # 팀통계 데이터 (이름과 같은 행 = row-5)
        stat_row = row - 5
        vals = [emp['name'], emp['dept'], emp['wd'], emp['p'], emp['l'], emp['a'], emp['lv'], emp['ot'], emp['lateRate']]
        colors_ = [None, None, None, GREEN_D, AMBER_D, RED_D, None, PURPLE_D, None]
        bolds_  = [True, False, False, True, True, True, False, True, False]
        for i, (v, fc, bd) in enumerate(zip(vals, colors_, bolds_)):
            c = ws.cell(stat_row, sc+i, v)
            c.font = font(bd, 9 if i < 2 else 10, (fc or 'FF000000')[2:])
            c.alignment = align('left' if i == 0 else 'center')
            c.border = thin()

        # 구분 빈행
        ws.row_dimensions[row].height = 8
        row += 1

    # 팀통계 합계행
    tot = {k: sum(r[k] for r in data) for k in ['wd','p','l','a','lv','ot']}
    tot_att = tot['p'] + tot['l']
    tot_late = f"{round(tot['l']/tot_att*100)}%" if tot_att > 0 else '0%'
    sum_vals = ['합계', '', tot['wd'], tot['p'], tot['l'], tot['a'], tot['lv'], tot['ot'], tot_late]
    sum_row = 3 + len(data) * 6 - 5  # 첫 직원 stat_row
    # 합계행은 마지막 직원 아래에
    sum_stat_row = 3 + (len(data)-1)*6
    for i, v in enumerate(sum_vals):
        c = ws.cell(sum_stat_row, sc+i, v)
        c.font = font(True, 10, 'FFFFFFFF')
        c.fill = fill(HEADER[2:])
        c.alignment = align('center')
        c.border = thin()

    # ── 시트2: 팀 통계 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('팀 통계')
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 16
    for col, w in zip(['C','D','E','F','G','H','I'], [8,10,8,8,8,8,8]):
        ws2.column_dimensions[col].width = w
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 20

    ws2.merge_cells('A1:I1')
    c = ws2.cell(1, 1, title)
    c.font = font(True, 13, HEADER[2:]); c.alignment = align('left')

    hdrs2 = ['이름','부서','평일수','정상출근','지각','결근','연차','특근','지각률']
    for i, h in enumerate(hdrs2):
        c = ws2.cell(2, i+1, h)
        c.font = font(True, 11, 'FFFFFFFF')
        c.fill = fill(HEADER[2:])
        c.alignment = align('center')
        c.border = thin('FFAAAAAA')

    for ri, emp in enumerate(data):
        r = ri + 3
        ws2.row_dimensions[r].height = 18
        vals2 = [emp['name'], emp['dept'], emp['wd'], emp['p'], emp['l'], emp['a'], emp['lv'], emp['ot'], emp['lateRate']]
        fcs2  = [None, None, None, GREEN_D, AMBER_D, RED_D, None, PURPLE_D, None]
        bds2  = [True, False, False, True, True, True, False, True, False]
        for i, (v, fc, bd) in enumerate(zip(vals2, fcs2, bds2)):
            c = ws2.cell(r, i+1, v)
            c.font = font(bd, 11, (fc or 'FF000000')[2:])
            c.alignment = align('left' if i==0 else 'center')
            c.border = thin()

    # 합계행
    sum_r2 = len(data) + 3
    ws2.row_dimensions[sum_r2].height = 18
    sum_v2 = ['합계', '', tot['wd'], tot['p'], tot['l'], tot['a'], tot['lv'], tot['ot'], tot_late]
    for i, v in enumerate(sum_v2):
        c = ws2.cell(sum_r2, i+1, v)
        c.font = font(True, 11, 'FFFFFFFF')
        c.fill = fill(HEADER[2:])
        c.alignment = align('center')
        c.border = thin()

    # ── 시트3~N: 직원별 일별 근태 ───────────────────────────────────────
    for emp_daily in daily:
        sname = emp_daily['name'][:10] + '_일별'
        wd = wb.create_sheet(sname)
        wd.column_dimensions['A'].width = 13
        wd.column_dimensions['B'].width = 9
        wd.column_dimensions['C'].width = 8
        wd.column_dimensions['D'].width = 12
        wd.column_dimensions['E'].width = 11
        wd.column_dimensions['F'].width = 24
        wd.row_dimensions[1].height = 22
        wd.row_dimensions[2].height = 6
        wd.row_dimensions[3].height = 20

        wd.merge_cells('A1:F1')
        c = wd.cell(1, 1, f"{emp_daily['name']}  |  부서: {emp_daily['dept']}  |  기간: {emp_daily['month']}")
        c.font = font(True, 13, HEADER[2:]); c.alignment = align('left')

        hdrs3 = ['날짜','구분','유형','시간','상태','현장메모']
        for i, h in enumerate(hdrs3):
            c = wd.cell(3, i+1, h)
            c.font = font(True, 11, 'FFFFFFFF')
            c.fill = fill(HEADER[2:])
            c.alignment = align('center')
            c.border = thin('FFAAAAAA')

        last_date = None
        bg_toggle = True

        for row_data in emp_daily['rows']:
            r = wd.max_row + 1
            wd.row_dimensions[r].height = 16
            date_val = row_data.get('date','')
            if date_val and date_val != last_date:
                last_date = date_val
                bg_toggle = not bg_toggle
            row_bg = GRAY_L[2:] if bg_toggle else 'FFFFFFFF'

            gubun = row_data.get('gubun','')
            status = row_data.get('status','')

            # 구분 색상
            if gubun == '현장': gfg, gbg = BLUE_D[2:], BLUE_L[2:]
            elif gubun == '특근': gfg, gbg = PURPLE_D[2:], PURPLE_L[2:]
            elif gubun == '결근': gfg, gbg = RED_D[2:], RED_L[2:]
            elif gubun in ('연차','반차','병가','예비군/민방위'): gfg, gbg = GREEN_D[2:], GREEN_L[2:]
            else: gfg, gbg = GREEN_D[2:], row_bg

            # 상태 색상
            if status in ('정상출근','정상퇴근'): sfg, sbg = GREEN_D[2:], GREEN_L[2:]
            elif status == '지각': sfg, sbg = AMBER_D[2:], AMBER_L[2:]
            elif status in ('결근','조기퇴근'): sfg, sbg = RED_D[2:], RED_L[2:]
            elif status in ('특근출근','특근퇴근'): sfg, sbg = PURPLE_D[2:], PURPLE_L[2:]
            else: sfg, sbg = '888888', row_bg

            row_vals = [
                row_data.get('date',''), gubun,
                row_data.get('type',''), row_data.get('time',''),
                status, row_data.get('memo','')
            ]
            for ci, v in enumerate(row_vals):
                c = wd.cell(r, ci+1, v)
                if ci == 1:
                    c.font = font(True, 10, gfg); c.fill = fill(gbg)
                    c.alignment = align('center')
                elif ci == 4:
                    c.font = font(True, 10, sfg); c.fill = fill(sbg)
                    c.alignment = align('center')
                elif ci == 5:
                    c.font = font(False, 11); c.fill = fill(row_bg)
                    c.alignment = align('left')
                else:
                    c.font = font(False, 11); c.fill = fill(row_bg)
                    c.alignment = align('center')
                c.border = thin()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            payload = json.loads(body)
            excel_bytes = make_excel(payload)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="team_att.xlsx"')
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(excel_bytes)
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
